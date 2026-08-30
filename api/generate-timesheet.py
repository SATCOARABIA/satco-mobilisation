# api/generate-timesheet.py
import json, base64, io, calendar, datetime
from http.server import BaseHTTPRequestHandler
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from collections import defaultdict

LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAXwAAAApCAYAAAA205I6AAA800lEQVR42u29eZxU1Zn//z7n3ltVXV3VO9DdgCC7CIiAgMqqRnHDfQkao5OoSTSZJGZxxoxLZrJqzKgxJppE4hJFY0QFFGRRZJV9b6BZe9/36lruPef3x71VXY24zIwzk+9v6nm9eNFVdevec0+d83me5/MsV2itNRnJSEYykpHPQTR8LKQKEOJ/dXRm5gfKSEYykpHPC+5PAuoe/v8vY30G8DOSkYxk5HO07dEfY9if5M8M4GckIxnJyP+rIlz7vrcS0EmY1+7//8uInwH8jGQkIxn5nEx8rVVvJSAALdBoEPJ/3cbPAH5GMpKR/ydFaQ0apHRBVCmXUEm+/h8fD+Bo4UG6a+1Lz8B3HI0QCsMw/ne9kP90lo7WSSflfy4aoTVaaxAC8XHX9MaF9ob19xApyUhGMvL/FyM+lYUjPiO2KK2Rfyc49ImA7wYhNEILtACptOuaSNnLMdHKA1mZ1HPS+47uYa80CI/DUlojhEADMu3ySrjHeA6Qq060RmvQUpCuGx1AK4WJREvtuVMakTa25I/jaI0hBEJoFMlriFSQRaRpaJl28647ptxh9HyC0O5YvRGmNHr6vInU+Omh8E5cON7H0rNUXPNEeGN0r5O6iE4uNJF+0Yxk5KT7Vpz0ddonOrne3V2ghMRd5Rq07Fl8Wnmvk/yzTm5U9xhvIWuRtGo9SkMZqbWrRO/rp04tkkajQKBQHgsuvHE43idSuztCuRsSCTgKth6uoTuWYMrIARhCsPFAFYYhmTisBEsq0Ia3X3Rqr0vvfrRIjhmkBp22F7WHTaTNo/aA2zwRuBU4Egygrj3Kq+/vYeSQ/khts/dgBZfPGkf//Gx2HKykO24zceRAgiYopHt32nFRR4jU1Kag1BuXSHchkh+mcMmdo8+F0nExxl0ZSghQGmlKSMTorm8hHk8QzMnBKsxFAHFtYyrtYZKB0AIcB2G43JXSCiGFO8GOgxDJG9WgleeKudcTQqEkaBsMywRH01VdSUdLE+FgLtmDBoMl0Y4Cx8YxTEwhIdJNV2UN0USCUFEh/n5FmELg2DaG4f7IQvcsPyU0OMpVQML1IDQGAoXQ0lU4RhpIoxBao4XhLhB69kFqzoT322jt7gntILVAC+nOjRIooRA4aExsnQbwSqGERnpnFiluMLld3J/4I7s6IxmB3kaM5w33gJ0H7lp7287AdmwEGsf0IYV2Ac/dGd53HJSSLiah3T2sBcpTFaZ294kykp87aA3SkUgp0EKhtXSNMTRIb18BSthoLRFID8AktgZDK1cBCIkCTJSLcZgu6glJ1Fbc9dQ6DtU3s+O3NxE2Teb9Yjk5oQDvP3wlBVkmSuMpDnevOdpA4SCVp9oMkGmIngR9F0a9OdDuLncckKZBN/D4XzfS2dXKD2+YTbbPh6McDGmw83A93/nNEq46fwoWNq+vWMfQIf3pl5/Nd55+jwO1rax54laG52XhaI0UCgPHs/wEGuWpTIHQGpRysUTi4ofyjFDlKrEkDrlGaPJD8Z8HfOmqNRLS1cpSKQ6+sZjNCxdiH6tG2wpyQpSMGcsZN19H0ZjRiIRC+AWJ5kYW//QJOqpqiGubM6+8jAlfvJruxgYW/exR4lU1mIZFQrrjl0qjpEJKA6fL5tTpk5nx9a+A5aN67YdsfuEvNO3dQ1dnJ76sMH1GjWLSbTcy8NzJ6ITAUA57X3+D3a8tJHG0ilgigVGUQ7+J4zlr3jwKx56GdjRCusrL0O5iliRctk0bKctZKI3ARkuf57YJhFauoe3OOIa3MYT3Je1ZOelT7mjXRtFCeEa6QqMxhEYKEMrzVIykonMtG0NphHDQ0l2BGoEjJcrdGhgoT2llED8jJ3HbP/a1BBRKKxwtMYTCtCwADA1KJ80J19pWSIRMGmUKrcGQLsDYnsHp2dDe2hQgPMNLuipDa+VasbZ0g5Zpg5MItEp62ybCcfBb0jurgeUdl9ASqRSGUmjpeh9KaDCyEDKUUgfS8KOsoGeMWS6AKxdEDa2RhoHCQHrOulLKPReAkC7eiaQ7rVx+Q2k0Css0aY8n+NGzK3n6ze3cduV44kCWBEu715g0qpSlj9zEkP790Gi+cvFIpo7sT9x20GYIIaPe7Rued6XQwkAhkWiUbSOE9nh+iVa2B/gWUoM0pcdaOAjlRgiETHoj8jOhwacGbbUQKDR+Kdj53Css+9d/w4h1EIjbKEcjLD+7P/iQ7Pwc+o45HQcwhKRy+2YO/eVF/AlFdzRCmWNzxjWXQmsrNcvehWO1ZEmTmHRIGBKhNQIHIQ0S7Q7hcBAsg7LXF7H8vh9DfQ2mKTExSIhGavYf5G/vr2baQz9g4nXX8t5Pfs2ep+eD3Y2WwrU8qhVl27dzaOUa5vzrgwy9cCYojSM8N04LTylK4tIk6tk2fkMQUAYi5QoKhKeJFRIpDdCuEujhdUQvykegkUKTQNAtJAYaS7tWlhAaMNDCIAE4nl6XEgwtMXSPe6yEtwk1WJ534V5AeBsjIxk5OdgnDY2UEStAKTAMAxOJA7y35yiRSJxpYweT4zNdG1MohDY9MHf3vyEgZmuOVLeQZRn0L8rBlDpFzwiP/jWQVDR20RqJ0784hxyfC2zCZ3lGEGnJizJFmwohcCyDD8ubWPJhObVtXZQWBLho4hCmDC1GYJAWNUTi4COCX3fjSxJBwkZoG5nymF062BCCbixeWrmXTfuqyAtnc93MUZwxqACNcvl1b8+iNY6QaEwM7RpowjAob41x/++XsGp7LY9/92ruOH+4pzwBLDSa/IDgvLFDqGrqoCPhMH3sUCygMW7jF3GCKuLtbRd/pEtlpGgtw/R5nH8cLUyk5U9Z746Gg7WtaCEY3ic3pQxd6NGeP/bp9I78VMBXGktIOior2PLsn8mOxTADAQIjR9Hv/C9gDTqVrqwARQMH4ClKAA5/uBlD2YRy/RTlZdFy4CDthyrAn0UiP49InzwifXOxwmFysrIJhrMxc3NQhXlEC3OhbxHV6z9k6b88hGxpQmYFiIZyMPsPRIZCGFIgWrsICIMtL7zIh0//noCwcXwBZJ9SxMD+KH+ALJ8fp6KKt//pQeq273QpJce11pWU2PjQ0uK3b2zhsh8+x5X/9DzfffIdOmzXzQLlWdm4P652uTyFxhYGjpA4QqC0a80orXGSxwqTFZsOcfUP5nPbgy9wsKoRQ5goZaHwKCEJjy14n/O+/Ue+9+TbdMQdpOHBvTZAGQhHuIoiqWC0SYpnykhGTkLs6N6bOAW2hpR0xh0WrCnjH595n2t/+ja3PbaMY20Rj99QIEx37Ts9dCJCcqwpys0PvcpDz68lAQjXHwUEhgOm4171ly+v5foHX+VQdSsmgqONEe75wwpefG+3a0AKjRYODpqElGgJMQ2/XLiFKx5YwJvr9tPc1s1ra6u4/oFX+eO7W3EMcKS7t5J+tI3G9tIeFZKYNNLiCAnXgBQCJQWP/HUD3/nNOxxp6uCVD45y/UML2Xi8wbV4bcejuxzPd/EoHe1gSsGmI/XMvW8BLy2vwio8haq2KGvL67CVRmonFSvUGjodh3/+0wqu/dHzHKxr8eZf4yhISBPHMwwNrVzWQClMoCkS5W/r9rK2rAItfLTFbN7cdJD3dh1CCEFzd5x7Hn2Df3liMZ0e7a+F9sZre/f86fk3n56WqVwaoan8AN0VlVhKER49losf+QW5xQNora6mbMNawhPGuBNlGdgNLVSt2Q5CEknYWIYk0dBI1ZpNjL7teq577N9xYt1YWT7K57/Mnuf/Clk+zrzzFgZdfjGJSIKscA5rfvZzZEMD2uen75lncta37yI8YACRtka2vPACdtygdMRIFtz+NfLtOJ15OUy86SuMmns5IlvQtHMvHzzxNPrwEaKVR9nw5FNc8fgjSH/A5fW0AClpjSteW3eEtXujYJlsO1LFzRe1cu6wQpTjIIXEEQZSCKRSCJEAaWGrnoCw4ZGBWmuUclz+UJqUtzqs2B0lL6y4qyvmxr6EQ8JzMyxMjlU0s3FHLSHDj7AV+DVxLbCEiSEcj9yXJBzD84pdy0pmMpAyclKw7yEb0dqlNQRIKdl0rIUHn1lGfXMb55xzJuH8QuLdnZ5BY5PQGqkU0nFpDaHcAKMSBq0O7G2H/Bho4iiliOPDBKQj0FLgKM3xSIJDnZqo43qhR5pjPPrqDq6fM5prZpyOoRVol65wtMQ0JBX1Lfz7ixsoGtiXxf92JQODFu8c6uCK7zzH4vWHueWCMzG1g3I00gRbQ1wGcaQNWmFql+tXKS9AoDQYhmDr8TYe/dsOLpkxnhe/fR5LdjZw+X2v8fy7e5n8DzNASxzlxthMHFchIdDKQEqBqRRThxUwfvgATBI8t2wLjy2IMv+7c5h71inEHccFUo8NqekSVLQlUJ6i1YAjLWLCIqbdOIhSCbSQaAekT3Ooro2b/m0hF547mr/98wCa27u567FlDCwO8u7DQ4g5cKjFIDcoSQBKKDQKQ7hBdqFlWkDiPxu0FeAYLuDb7Z1exosiVFBI4fBhYBkUFQxn2pjhLkNtK6RpULdnD5E9hwgIPwNGjaKhsR6OVHBsw1pG33Itp5w5JnWNupFb0QmFDEDBkKH0P2McAA0btlO/aQs+IZB9Spjz0wcJjxkFCgrlqZSeeTrxtghbfv179PFaHNPk9BuuZdr996R82MIRw8npV8Kir3+LrKYYdeu20HTwAIXjzkQ4DmgHQ1qUVTRRXtuJCIWRhkFXtJsN+45x7rBCdzFIaHVc7jPPNHCwULZDUGhMw0Ij6LATKK0xpCBkWCkby/AJZCgLKySwggESKKKOTdgKoBx303zzppnMvWwSfYMBQn4T7TgEDBMb6LAdHK2xTEHQlK7noBIut5qhdDLyqdk6yUQE1xj5xZ8Ws+JgghWP3cyE4iAf7jrM8S6H3CwTKSz8RrrvbwMSn0cEZAdNgj5JtiUIGCYSyEqSBEYPoAR92fgNsyfxzJD4gya5WRCQ6VRk8n+H4SX5fPnKM/n9kh08//5BrjprMPMXrqNP2OQbV55NlhTu2Q3X+PH7DHzSQNoJlHQ5fVMbrneCe6wjFAZQdrSZ1i7FeRP7I4HxA4MUFQQ5eKTBzbqxRNp4jB7awwPQSUOL+dN3LkvN69LDTcz5wWu8ve0IV04ZTED2UE05QCDgw5I92Uim1phOHJ9QhAMWUkikZbmT7CFwMMtHdnaQQDCIIQQ5QYNQMAt/MEhQCOyAhS/oR1lurK8nq8nsyeb7DIkcn87he6GZQEFfbNNHjs+hZt16/vb1u5lww3WUTjwDX0EhOu7xE8DhzZtw2lqROSHG3XIz65e8TfvhI1Ts/pD2w4fJGTmceCKG3/ITj0XBSLg8upNAKZduKd+4BtXWilIGoy86j/CYUdgJG0O5wSDDH8SOt1C2+n20NDBKBzL5izejhcax4whhorWgeNokBsycyqEFb6Bbo9QePEbhuDN78vWBDw/W0djcRb/SMNnhXA7tL2d9WT3dl2qyTIvGzgh3P/4mZRWtXH3+eA7XRqg4fozffXcuhi+LpxZuZmNZLV0xm1y/w9SxA/ny5WcxMi8by46j7ARxnc1f1x3np39YSVt3N1+69FxumzUcqeHN93bw+obDnDWihJ/8w0XkZ1ms2lvLn5dt50BFA91K0CfsY+65I7npgnHkCYlQOoP3GflYHl+l7eBkmrMSglNKi1C7y3ngmZVYTpyd5fUow8frm2uZdEo+dsLBFiadnW0Eg1nYWEQ6IgzMz2JzVQdtEUVdU4ylZS34RTJcq4lEQTkxDL+fHQdqyRbd5AS8DBKlMLWmsk3y1u56/CiEZdLS1kGfnDAJJ0FbzGbqhNG8Vxblvic/4LGiTTS1dnHTFVOR4TDLdtViWga2StAW08RsOHq8ir65PizLT8xRJBzIcpwTsEsyvG8WYZ/Nyx8cYUhJH+oau9EONEUUSw+24FcaJxGnLRohPxTGUa6Fb1kmHZ2dgCCQFUApjd8QHG2Kk+UPcrAhzvpDrSTiivZ4AkfH0dJky5EWQtlhQgEfGghbJiUF2azeW81f11QwbmA+StmYpkFLWwf5OSF217TRGjOobEqw6lAbR2paqWoFMwgry5upquvkaF0HEwYFyDY0QuFmOQIIx01WFZ+eoPmpgG8g0UrTd9xoCqdPpPatpWTLAFWLl1Cz6gMCo4Yz9YvXc/r11yKkheps5/CGdTgyhr9kCAXnnkP/yirq3llOrK6Jym1bGXPaSEwtEFJiCBO04y4cIZFSgtJ01dUg7QS2ZZE7ZhTJymTbVKDA0Jr6XWW0V1USdKDPuDPIGT4MRytM2cNoaSEpGTea3QuX4MOhtbLGtSmEwBCCuNas29eME9OcPaoPpcX5PLlvH9uPtlLdHGFoUTZRW7PlaJzyaoOKxeU0V3UycJCfiJXFPU+8x4pVeyGcjz9LYNsOq3Zto77b5A93TPNqD/x0xvz8fsFGhLbp0gE2HVtLYTibayaWsruyk827Y8TMKNoy2FbZyo0/eYv6RgVBE8unSMS6+WDnGvxZYW6fMQTHtpFa43gLPL2CL1kQor16h14K/IT30l/rZGEbrvuffM9xXC5UStnr2JNdK/39j/s7/dpKqdTnH3f+1N9eslz6LSXPkRxv+veV6qlsTJ4jeY/J8aZf88R5UspNyzMMIzWO9DlKfj+Z8ZE+dx83T/9TYdue0g+BkC4dqYF7b7uIoYOK2bTnCHk5eZwzYRTL1+/l188son9hiHjCpj3hELXjWKaFxocTjVPod2izJSMGlOLE2vnuL15ASz9CuDGueMJEqxjSBMufzb03ncuI4oA7Lyh8lsnGXXXsLVuIRRwMSaQ7Qa7PRwKI2DYFPh8YFlNP70MwJ0hjfQ2bNn7IhjXrEXYCS0psYdNhW/gxGJQf5NtfmkZhwKKuPUKp1UrQl+Xx7zYWCkdrzhpVzAP/MJPHFm7i9p8ewFI+InY+h+q6uedXr2IkbOKOotNOEPKb2NrN+TelQSLWjUBg+iwve8+NrRUCh8u7+MYvj5KIC2LKARUhIKFfOJ9vzJtBaV4I5ShMQ/K1q6dQ1xHnsReWETBcSss0JF3ROH6fn25bMGpIX6KRDr72s+eJxzT9+xThlzbf/ul8YrbBuNJCvn3dOWSbEm0LpOn9yqInteS/xuFrd6EqR+HLyeXSB/6J90NhDq1ajdXchtUZx964i+U7d6GkZOy8G6jfdZBo2UEEipLTRhA6tYRBE8axJzeXREcLR9d+yOnX3YAwvEsLM5Wx0hNDVui44/oWEiyf3027Usl0Lze6XbltJ76OCIYWDBg/Cm25HKKb5KRTBV+GsDCRSMdGq3gqk0YakqONEbYdqgVLcuaQABOG5/LH13wcb4iy6WANQ4uGIQBfVhYyaBBLOFx53hDmnjWAIp9kQIHJlPF9ueCsU7jpiinc+/Ra3lp5mHV762mJa0zppns63Z3MmzOGK2eO4J7H32P/sRgvrdzPZRNLCfr9yECAkN/CQGN3R5g8PJ+m4m6+dv3ZDB5cwu0/WcKBQ82s23mc22cMcVNAhcA0e37CE0EnCVrJ10lwTQe+dCA7EcCS3zvxGr1CPGmAna540sfSK0vAG0fy7xNLzZPKJTme9DF7lLRb0Jf23onXFF4ltmEYvQA6ea70a9q2nQL0E+8x/bj0MZ94/ye+n64kxSdVhf83gn5qNMnaD2+z9/UJ7pozHuaMTx399QuGUtfchi29tD/HQRuu4SWFm/AnDIktoDAnDI5NU0cXQkNCuTk3lnDz1BNK0Tcnh+JsPwZxQGIHAnQk4sybPoB7rhuH49juGJNrQbjWqnIcLDRFhQVkBX00NbXR3hVBJMflKftkDUBxbpj+IR9aR8kN+Xj2oesRUpJlmaDiaOEDIZCOzT2Xjefm806jszvCvvouvvjj95k0ophffXWiF3h1UxxT61u4v6MUspfi7plW6S1GD2eElxuvFH1yw5SEs706CIVWmhnD+/DKv1xBRVMLWjk9v5Ps2X+5OWGkUrS3d4AUhEJhNxDc1YEUgv4FeRRk+bG1RhrJugr3lxXwmbK0zU+1FbSDMATaVuQNHsbcRx+mdssOyle9z4HF7yKPV2BHu9mx4C1Ov/oaqrbvIlbXgj8QpLO6ivW/+hXdVTXILE2gS1O5aRttxyrIHTYo5XYpoTHoSfFCGvhz8wALFXeIVFS7ARGlEMrA1JJEZ5TK7dtRONj5OZRMHO0mJTkabbrFTdKrsmspP0owZqMNH/nFJSk3Ewy2H66nor4VIxwkv6CAPn3z6du3iOPHW1i7r5Ibzx6GiQKdQMWijBxSyB++dwGFprsQ/vTNC9la28S+mih/erOMXQea0IEA0ViElq4o2jBAxcjyK66eNoiLRxXz3NAwB462U9nQRlPCcfk4HccQDqo7wVnDS3nx/qvZerSByoZu/vzXzdS2xBG+AJH2TmKOg98wOXb8OAtff52cnByuvvpqwuFwL/BLByHHcVKWZxL40sE9CY6bNm1CKcXkyZMBaG9v55VXXkEpxXXXXUdubm4KxBzH+QhgHz9+HMMw6N+//0fj/0r1srpt2+bgwYP06dOHVatWcd5551FYWNgr+G0YJygb4eVtpymEpUuXMmnSJPr06ZMaU0VFBfv372fWrFmYppmaDykly5YtY/v27Vx66aWcfvrpH93M3r298cYbHD9+nBtvvJF+/fqhtWbXrl0sX76ceDzO7NmzU/NUXl7O66+/zmmnncacOXN6eUj/sxb+p1C02u05I9AgHNCKoixB0YA+H4MYLqh8eKSWx1/bzPkzxtNUXcWx4y386Mvn0y/XSh3zx5XlrN52iO/dMInioI8tFVHKG5t5YdVuTC24cfpwxpUUfuaB5hXmQmHuJxzkgBNHIAgAo4uL3DtwHLTwYyO9mhbB/ooGdla1E87P4eU15UQiTcybOZEz+xd97nETt37SrakRQqK1wHY0hT6DwtLPcL28UO/X+dmeGYwL9r38uP+YfErQVqGlx/+ZbiGEkBYl506i5NxJjL54Novv+kfU4Ura6+tpKj/M4c2bMBwHn5lN0/YyKjdvRRoGIZ+Bz2cRqamjZvtW8jzAxyvxMJX2SqhdKRp0KoYvgFAOB95bxaRbb0QW5GPbDpZpULt9O41b92IIi+DIYYTy8klEIljBIMpRKKUxLUnXsQqOrV3nVtrlhSkYOcLbgO5G3LivlljUwQpl8+hzG8nyOzR1WRAw2HiwmS5bIQ0TJS1QnQwbkE9YCmzHwUbw+FtbeHH5Lo7WR+nXtxB/VhYQR2oHB+WlYUmyfAGyTQPHUYTDJtp0cLRNPOGmYSJM4trCsAyOd8b5l+fWsGbLUZrbuhkyuB+hrGzaWzuwvRS5luZm7v3hD8nOzqaqqor9+/fz85//nKbmJgxpkJeXR0dHRwpwcnJyiMfjNDc3U1xcTCQSIRqNEgwGkVLS2NRIaUkpS5cuJR6PM2XKFLTWPPzww5SXl9PZ2UlZWRmPPvooXV1dxGIxCgoKiEQi2LaN4zjk5uby+OOPM3r0aObNm0dnZyfhcBgpJU1NTfTr1y/1d25uLhUVFTzwwAPcd999VFdX09HRgd/vJx6PEwxmEQhk0dHR7mY8CUEoFPIogh5vwXEcnn76afr160ffvn1TiuDAgQO88MILzJgxI6XwLMti0aJFPPnkk5SUlLB8+XL+9Kc/MWDAgI9QS4sXL+Z3v/sdoVA2mzZtYv78+ZimybJly3jnnXcYO3YsP/3pT3n22WdxHIcf/ehH5OXlsWjRIrq7u7nuuutO6j38PRD8wkh6AAYogXY0OlnxiXa9aK+CUzsKaUgOVTXy7ppd9B9QyuHyoxw4UkVT5Gz6hnOwEwmEz2LXoeMsWbOVm84fzqjSfP70ynIWbztKSWEOv7pzJueNH4CjnE+kHrTHKiTdc32CAeoWlGo0CZAGiIAXs9Box8FAID1O2xDSqw4zWFdew7/9eSWGoQllZ/Hw7TO5ecYwlIqi8JFqK/HpmeqfrE1TnrKnXLVroBheSxmtVBrZlixEEynKMuV9C+1V36bBujghEUf0VjSfRQWYn+YYag1CCjpq66hau4VhUyeQ1b8PYBDKzUP6JYajyJKS9sMHiO7Yg99y049kKJssI4ijQcUT2NqBrm4OrV3LyCsuR1qm12LA9dWSvS0SaPqfOwmrfzFUV1O/fRsf/OpRzv7ut/Dn59FxrJINjz+Jv60dZRucPnMGG559lUhzBxfe9x1CQwcgLYPuhkbW/OLfaT9SjmM6FEwYTd8RI9FKIw2L1kSCTfuPgeHDUQZH6rwpy7Ldoq+qLnZXtDC8JIxWJiibIl8CS7i86Kp91TywYCvRiORHN03iB9dP5t4/r6bsYDVChzFwK4gRJtFoFNtWGIYk2hpBxMEyTII+w7VctYOhNI5l8tLynTz35l4K+mbz9H1zmXbmQK7+4V+prkoQN3z4pWDh8uUEg9n84Q9/oKuri4MHyzlw4AA//vGPsSyLH/7whzz99NPs3buXYDCbX/7yF7z55pssW7aMb3zjGxw5coSXX36Zm2++mYqKCtatXccdd95BOJxDZ2cHAEeOHGHfvn08/fTTaK3Zs2cPhw8f5t5776Wuro6HH36YPXv28Nxzz2NZJtdccw1bt26lvLycgwfLef/99/ja177GgQMHWLFiBd/97ncpKSnhvvvuY9iw4Zx22ii2bNnCW2+9hVKKdevW8fLLL9PW1sall17MZZddxj3f+x5SGIwaNYpf/epXvTyTpIRCoV60E4BpmoRCodQGktKNRS1ZsoSvfvWrXHPNNSxdujR1zIkxgMrKSvx+Pz/4wQ/Zt28ftm1jmiZSSubMmcONN97InXfeSSwW44MPPiAvL4/f/e53bN68Gb/ff1L65+8G81OBEDcVURk9vZ2E0hjCSycUbpW4o+Dqc05n9p8HkRcK0m2Poct2KMnOQmg3U00om598+Vy+d91EinLCKAT/9A/n8z1HkRcOkh/wu62Dk+1UPkak9mzZVBOuE2JQhpsfL7yCRCV7emC5ue0uUCrhUqmG1KBsrjlnFDPOGIx2HHLDQfoE/GgFtjS9qnvltj5J8l+pJow6VW2vhdcqBa9gMy3xw43VuddX9H7yldbJTlwubmivj4MQ6WCdbOfgGaNJPvoE7yGpBDQfjc/91yttNYgECL+gs6yMRff8gAHDBlM6+UzMvELq3t9I5EgNpiMpHFhCV301nVWVCCkZMvdCxl1zNXY0hrZ82A0tvPvoE3D8KNVbt9FVVUt48AAv5mB4/TS8CXJscocOYeR1l7Pp4cfJs7LY/sLLVG7fTcGpp9K4s4zYwYNIZeMfM5zsgjzWz38ZGttZsG83AyaNxcoOUL17P53b9+AXgrasAJNvvg4jlEXC8xJ2Hmlmb00c4fNTGOikINdd6F22pjbhp6uzi817Kjit/xgkAin8mFYw1UKhtrELW2XhD5kMHlDIjiPtrNhUj/Rlu83avIUhTE3EUSzcfAwjHGLbcRst/QwtzSXPNOjGQmJiAbZS1NQ1YfhD5IdDFPfN5Z33D1F2vA0ZCGB6Qdqq6iqGDRtCY2MDjz/+OJMmTWLr1q0crzhOPBbllVdfoaG+niuvmMuGjRt58S8vsm7dOrRSPPPM0/Tp24dzzz2Hyy67jAULXqKoTyHvLH2H8Weckerp3dTURCgUQkpJXl4eM2bM4MEHH2Ts2LFceeWVPP/8c4RCIc44Yyx5eXns3buHKVMmM3LUKLZu3crMmTOZPXs2x44dpbAwnyVLFiOE4MtfvoWS4hJsx2bq1CnMnDmDZ555htzcXEzT5JZbvsTSpe9QX1/PlMmTKSktZeXKVallWVtbSyKRoLS0NEXrOI6D4zgp2khrx62HUA6O41JQ8USMeDyWon4uuuii1Ibs6OigubmZwsJCwuEwN980j927d3Hn1+7k9q9+Fenxu8FggCeeeJLFixcze/Zs+vbtS1VVFYMGDcK2bSZNmtSLGvp7o3R6pfC53KbXmMCDHSmxtYGhFBDHMlzlZWtNcW4YlE3A7yPf71aIux6Vg0YTsCwG5Pm9PRxnQEFOCoYS2nHrWD4FlpRIs3VF73HrFLUn3EYOQriALtysFbSBI11DVQqNdmy0dtBakOMzyfHnpKgRx3a9Gq1x+1YJjdc8BZHs/6ac3srIML3fsyfN1QV9N51ASOG2StBe00e8po2pqxo9lJrXN1k5NskWk8Iweix+kTw/bv8i4XbXUo4NHjX5USX+XwV8wDHcg5r27Ce7pYP23Xtp2bEHHIlPWkgBkaxshl9+Cfs3rqUtFsUKhRl19ZUUz57ZqwNln6XvcbS8nK7DFdRs3Ep48AB0QhN1FAmlUF6lnmm71v6kr95Cze4yKpYsJxiwiGzeRcfGXUjTBFsRCWdx6ffvJlxQhDQNlIhiVxzlUPlBHENhSQuhDDoMgyl3f43hF30BbWu3ERGwfm8dTQ2tEPDx7evP4UvThyKEYMOBFm791TK6uqO8t6OeS8+1UXYMFYthO0k9bDOyT4igEaM9Guebj72LT2nMQC5KSNoSCeK4FXa6K44RLuRPSw4yf/FuOuwcfEG4/pwhWEDctlFxm3hckS1heP8CHGcfh6oTXHTPK/joRgcKUB0xulUCBxh66mDeWrQIyzJJJOI89vhjDDn1VAYNGsiECRMpKSnm+PGjTDprEkePH6OutgYh4IIvXEAoHGbrli1MmjSRrCw/ZWVl5ObmEAwGMU0jZXnk5+fT2tpKIpGgoaGBvXv30tbWxuzZsxk3bgzLVyzDH/BzxvgzCAQCbFi/HtM0KC0p5nBuDmeccQZK2Rw5cpicnByysgK0t7cz/szxTDhzIjt2bscf8FFSWkwg4EdIGDrsVCZOnMCKFcupb2hg6tnncMrAAaz54IOU1fyzn/2Muro6fvvb35Kbm4thGKn/k/EBKQV+v49AICstuGsipaC2tgbDMFi0aBFTp06lqKiIpUuX8thjj/Hggw9ywQUXsHLVSubN+yJ33HE7t99+O9NnTGfc2DOIdEe49tpr+OY3v0VeXh5SSgoKClizZg2mabJx40b8fj/jx49PxQ3+nigdt7+M9ppfCrf4UPc0FFTCrWiV2EjDz8qV7xLKKWTypAlEowl8lonw+upoKXG8gKujBT4heX/delAOM6dNoysRx2dYmGhMqRAYXtVv786yqUC710Fep/XW1F7bhWTHHSXA8Lw5twGZx5p7XLnSbuNFEwWGmYI4Rym30aJ2iBsSvwRT9nymJEiVQHgUDIKPeI1aO2n0iuhpK4B2A9Ha8JJUFZZ0vYWkgnC01xZNgGkm4zuq1zWSlr9OJdUKhBSYSQRVNobX+0h5v2Ga6f+ZEP/TK229rJjCM0bTZ/Y5NO4vQ7R1IqWDyDIQg4dwwV23c+qEM1n/0qv4BpRQOGQYuUNHohyFrd3qN79hcur0qVRtXYuDw/HyckYAZk4IcUpfCAUgHPCUjNstLlDYh7kP/4w1Q4dy8M1F2PUNaBVDB4NYI4Yz5xt3MuKSS9FScPXvfs2qf/81zVv3uQ8bUAm0z48x9FRm3nITZ958A45pIhyNm/mpaK6pZnQ/B39YcsHp/RkYdoMj04cbTB1icTjbRzTaQGdnB6f3g0RnjP45CYTHCU8Y2Y9//fI5zF+6DSfaxYVTz2D0qaX8asFKwkaCeFeEvkHJaUUxBg3yMWPKdN5ZtYGuaIwvXTKFiyacggaG5mpOHWAztAiwHW6YeToHjrewcns5VlaQL188ja62Dv701kZyfDFaY3FmzZrJyy+/zPe+/32kEIwdO4azJp3F3/72Gnv37GHsmDHEojE6OzpoaGhgxPDhaDTbd+xg2rnn4vf76ezspKmpiT179zBo0GCklHR0dmLbCQCGDh3KmDFjuOuuu+jo6GDixInMmjWL3/72t5im5LLLLqOmtoampibCoTDd3d3k5efz/PPPg4DTTjvNiy+UUVLaHyEEY8aO4YH77ycQzOKmL86job6Bl196me5oN9HubtrbOujs7KS1tZUrr7yC5/48n2B2Nn6fPxW4feihh1BKkZeXh1KKhoYG7r//fgYNGsTQoUMJh8OEc1zu/Vvfupu+fftRXNyPAQMGct111/HII4+yePESamtrGTfOLfS75JJLmDVrFtnZ7hpobGzkl7/4JWdOOJMRI0ZQU13NgQMHiMfjmKZJYWEhu3fvZt26dcydO5e3336b22+/nf3793P33Xczfvx4/rOPmvjvDdr2BPOVdlwQFj30vRAaQ9loIVm3cSM//+WjFObn8sjDj9B/wACU15ZECNuzWEWygSU7d27j0V/8FCENivsUMXLkKGyvL4lQ0ms8qHtx1SKlED1vIQ22BD21PUmvUyJ4442F2MrhqiuvSrUMVwK01BjK7cfV3Z3gxReeBTS5eXlcdfXVrpduWW77ca1Y9s5ibMfhokvmYjuKLMPtZePYCQzDYP78Zxk5ciTtbe3EYlHmzp2Lo1xPJdYdo7m5GQ2EskPk5uWiHI1puHfQWF/Ptm1bmDZ9OoGsbAxPuWitWPDyAvoV92PWrJm8+JcX6I50Y1omN97wRfx+f1pmkKKttZWX/vIXLrnkEgYNOoWnnvoNU6acw4QJE9zAcG/U/1TQ/+QHoHg95h0UpjSIt7fRsG07jcerSSiH/OJi+o09nVBpCfGuLhKVNW4702CQQL9+YLgNyjQKJDidUWJVlUipITuPQGkxTnMjicYGHCEJ9C3ByMvBdlvuI2yNYUqEdmjau5+afQeJdbRRVFpK3zPHk9W3iIRyXTG/hER7M5Wb99JYWY2Md5NX1JfSiWeQNbAUpZKuUaoBKc1xm5hy+18XBfz4lNumWRuShrhNXCksrcn3++lIJIg7iqDPJJzsMOhp1cZYAqEc8rMCaKAx0o3UmpDH5bbFYpiGJD8QoCORQDsO+YEAyVbQrbZNl22TJQ1y/JbbzhloiESxLINCy8J2HBqjMaSQ5PktfIZBXV0tr7++EMtnMffyy+nTpy8vv/wSSimuvPIq1q5dw5gxY9i3bx+DBw8mkUiwaPEibrj+RiorKwiFQowZM5bXF75ONBplQP8BmKaB3+9nwoSJLr3V1cVrr72GUoprrrmGcDjMm2++SWNjA1+65RZ27tiBZVn4fD4aGxspLS1l3fr1FBYUMHz4cIYOHcZf//oqhmFQVFTEGePH87e//Y2ioiIuv+wyli9fTnd3N+FwmPz8fNrb2xkxYgSbN29m1KhRvPrqX9m5cwfZoRDPPP3MR7KPtNa89dZbHD58GCklp556Kj6fj5GjRrJ69WoaGxrJyQnTt29fcnJymTVrFitWLGfXrj1cfPEcRo4cedJ6BaUU77zzNmVlZVx11VUIIdi7dy8DBw4kFoszceJEjh49yubNm7nuuus4duwYb7zxBqNGjeILX/jC3x2dkwxKCyFobKxDGgYF+UVuNpRwu0KC22gMBcKQ7N9/gLq6GoRwO022tnVy/vnnE/D5cFQCvN482tEI0+TosSMcLj+I3+fHtCwa6huZdf75ZGcHwStm6ulG6QbiY7EYNTU19B8wAMvLpkpPKz527BjFxcUEAgGUF0B+/7330EIza+Zs9z0pexpKaBuERSJhs27taiorK7EV3HLLLUghqK6qRlsG/fv2494f/oB4tJtHH3sCgKNHyhGGxaBT3ISSCy+8iMvnXk5VZRXtHe389sknSSRsLMvkyNGjPPmbJwGYNm0aF174BSzL4uiRwziOTSQS4Xvfu4eXXl5Av34lVFRWkJ9XQCiUzY03fpGxY8dw3333sWzZu+w/sJ8hQ05lzkUXYxiSmpoapJT069ePY8ePce011/DrX/+aadOmM2XKZO6++26+9KVb3Iw0aaTn0H8Wjf8JorRWdlwrx9a2o/TnKx9zPqW8/2yttdb2Jx2b/Ipjf/arKqUdx9bKcT6fu1Cq19/qMx7rKPWJM5P+6cmOVerz/j3+4/f73y1btmzRt956q547d65e+s5Sdz3YtnYc5z89jhO/p1LrTX1u5/x7E6WUtm13j6xatUpffMkcfc01V+ndu3dpx3F0IpHodXwkEtGPPPKIrqqq0suWLdOrV6/W7777rp49+zzd2tqqnRP2TjwW10899ZTev/+A3rhxo162bJnetHmTnnzWZF1dVXXSsSildH19vb777rv1jBkz9H333aej0ahOJBLatm0diUT0Qw/9WE+fPl1//etf142NjanvvfHGG/r22+/QLS0t7v5IG4+jHHd/a0e//PJf9MWXzNHz5t2oE/GYfu21V/VFF31BX3zJHL1v3x7981/8TN977w91NBrVv/zlz/VVV12hr7jicv3vjz2qE4m4vuaaq/Uzz/xeP/TQg/r7378nda01a9bogwcP6ra2Nt3S0qIjkYj+7ne/o5cufUc/99yf9a233qLXb1inzztvlm5orNevvPKynjFjmr79q/+gm5oa9Z1fu0P/6pGHtePY+ue/+JmeM+dC/aN/uU93d0f0m2++qS+88EI9Z84cvWnTJl1bW6dnz56t16xZqx3H0RdccIFesGDBR+77s4r8tBQuBwMHSUJp9lXWUVZTz9G2Tg7X1nOsrpFNB45wqL6Fow0tlNU0UtfcQkNLCwdqm9hdWceeynrW7D9CWV0j++qa+bDsCF3RKI5yOFDbyM7KBjqAbRW1HG9upS0ep6yqnpaooqyiluZIlCP1LeytbmTrsRr2V9fR1NnNvuPVHKxvZn9VLXFHsbeynp0V9WwsO8SB6jr21DSyr6aRLUcr2XW8mubuGLuOV9MVd7wuf96TutIKcz6aYaXTKjRJ9SQ5MWc7+ZlI9r0/4Zza66SXfmx64zOddkzSKRNp78lexVK9A4LJYGXyesnX6YVIyf+Tx6e/l/6dZMAzvbjqxGukB0lTufXeuZPnTJ7rZOdP/zw9Nz/9X/KYCRMm8MQTT/DSSy9x4UUX9uLETyyU+mjQtmfc6e+njz+9qOtkBVLJ43Ta/aXPz4mVwun39fdF47jVwtFolEceeZjbbr2V0aNH89RTT6UK65YvX878+fPdAGsiwUsvvUR9fT3Lly9n2bJl5OXmpbw/KSXr16/nj3/8oxvNcmzeeOMNKisr2LBhA2+//TbBrCCmZeJ48/HMM89w5513sm3bttRv+O6771JeXs5vfvMblq9Ywc6dOzFNE8Mw2L59G6tWLeffH/s1lZXHWfbuslSK68yZM4nHYzzwwAPE424/qqeeeopv/eO32LljO1IazH/2WV588UVmz5pFJBIh0t3NH/74R77+9a9z1qRJ/OEPfyARj7veB4o33nyTO+64g2uvvZbFixZhmhamZWLbNhrd63d95JFH2Lp1Kzk5OeTl5ZGVlcXRo8eora3ltNNOo6Wlhe5IhFAoxOFDh/j975/mJz/5CV2RCAsXvo7PZxEKh6muqWbJksU88MADbNm8mY0bP+Qvf/kLX/nKV5g+fTrPPfcc2dlu2nQwmOXNG73W3+eah58qHxeKpoRiwcptjBlczPZjzUwd0QcMi4UrtnH7tdPZtvswffoWUlFZQ14oRF5RITv2HeTc8aN4892NXDh7PHsONWPHI4w+tT+GNHnjvT1sLzvG9VdOZ+2uI4w7JQ8noWjrchg3XLH8/XXccMVsVm3YRV6fvixauYVzTy/lrDNGsGHjLgqKi4l3tjDw8lksWLKac6eOZ/n7W5k16TT213VRWlLE2s3bmTZhFEdrWthZVs5Xr7qAkD8b7bhP3/o4N0ickCYgPqGU7aPHio/M4Sed5+OeC/xJ50m+PrHwKf31iamG6Z+f7L2P+/1P/DydUkn/O3nOTxrTx43xZNdVSqXSJtMX+Ynz8HHn+Lj7+qzpkh/JhDjJ/KcrjP/tB1R/mtgJm7a2Nrq6ulKZSRs3bCCQlcVrr71Gbm5uSsnl5OSwc+dOampqyMrKYt36dfh8Pt5//3369OnDwoULaW9v57bbbsNxHMLhMPv27ePo0aM0Njaydu1aHMchK8sNmp911lmccsoplJaWpsazb98+QqEQY8eOJT8vj/Lycpqbm5kxYzrl5QdRSjH6tNMoLS3hyJHDvPfeSvr3H0h3dzdz5szhwQcf5K677mLEiBFMmTKFYcOG0a9fMQB79uwlLy8f21ZUVFSyYuVKmptaGDJkGK2tbaxbv578gkK2bNnC3r1ljBs7jvvvfwDDMPjCF77gZrwdO044nENFRRXxeKLXutqwYQPDhw8nGo0C0NDQQCJh09bWTl19A+vWrSeRcIjHbTq7uujs7EJraG/vxGf52bp1K8FgkO5IlMamZhIJh7Vr11JbW8upp56K1pr33nuP1atX09TUxNatW4lGozQ0NBIMBv/Ta8B48MEHH/ykvEyhbIQ02FvRSGO3ZGRpHnvKazllQDFK+tlVXsv0M4dwvLaD00YNory6m1jUoX9JEUg/55w5hFhHlGmTx/Hqqh2MGVLC5BED6IzbLHp/J4MGl9Dd2UVun3xkwsHvD9KeUAwZlE91TQslpX1oam6l3ymlbNpXwxlD+xKNWxSX9KErFmf6lDNoaG6hsyPCzLPH09IeY9KYIRyrrGXE0FLe//Aws86egE9HqWtNMGXsUK9jn1cenZG/W+mtKMVJlUtGPruV7w/4KSgo4O2332bt2rUMGjSYxoZGamtrKSsr48ILL+T0008nEAgwf/58NmzYQH5+PvF4nH379vGNb3yDvXv3cujQIQ4fPszUqVOZNGkSPp+PhQsXsnz58hTAb9++nVtvvZWzzz4brTUlJSUMGzaMUCiEEIInnniCJUuWkJ2dzYIFC8jNzWXatGm88uoCjlcc57W/vko4HGbRokW0trZy3nnn8eZbb6IcxdNPP8327du56KKLuPjii7Esi5KSklTAHmDEiJF8+OEm4vE4RUV9WL16NdnZ2WzevIXVq1cz74vzuPjii1mzZg3vvPMO0WiMU04ZTE5ODjU1daxYsYLhw0Zw+x13UFlZSV5eHpMnT0ZKySmnnMLq1at55513WLduHatWreL000/nrLPOYv78+Ti2Q21tHVdddTWXXnoJaFi0aBFbtmzl7LPP5oorrmDJkiXs2rWLUCjMli1bsCyL6upqTNNkx44drFu3DiklO3fupLS0lLKyMlauXMnkyZOZN29eap7/o/vg07N0vEKDQ0eP47MT1DXEueicEZimZPOO/dx4yXg62iLEohGaqmsYNSCHgqDF4aoK8nNz2VN2FMe2OXqoimi8C0UcDew4cJT8sMk3bpjBtu0HWLOjnLPOmUBFdTV+J0ZnSxtBK0H5sWokNlUVDeQHFaGgj7r6Gvr1KWD//jIKwgZtHREK8nPYc+AYTqKTY9V1mCJO+cFDZPkk3ZEIsY4oPuL0PC9ceZlMGdD/ewf7z/J+Rj4b6M+YOYOiokKefPI3FBcX88/33Ydt25SXl/P888/Tr18/amtraWtr45lnnmHSpEm9muedd955mKbJv/7rv/LKK68wdOhQuru7OXz4MD/5yU+46KKLPhJ4TVJe6YHjt99+m9tuu40bb7yRFStWMGbMGAYPHsy5087l+9+7hylTzub+++9n5cqVDDxlIKNPG82FF36B7GCYW275Mlrr3imNJzS1Gzx4ME899VSPd+NVPSct5alTp2IYBs8995xL22iN5SVj2Lbdy2O78847e41/6tSpTJ06NXVccixaa6ZMmZJq4ZG0xC+77DIGDx5MfX09hYWFjB07lldeeeUj6zlZILh161a6u7s5++yzex2TpOY+ztP9bHUYn5ilo0E5OMIgqhy3Yb/W+EzDfWh5IkGWZWErhZ1sooTGMg26bAeEdAs4hMDRmpjbvZlsQxBzFMIw8LlJUnQ7mqBlEk0kEIaJgSKhcHvROG4jtW6tCVsGOMp9zovjNibymxJLCDe7B7xcXLfcOgFY3kOBhSHxS4lUbumyFJnnwmbk/wbQJ3sYrVnzAQ8//EtGjBzJ9757D/283lL19fX88Y9/ZMeOHQghuPjii5k3b14vWjA9htLS0sKzzz7Lli1bcByHmTNn8uUvf5lAIJA6PtnX6MTOoVpr5s6dy5e+9CVuuOGG1DiTx3/zm9+kf/9S7r33n3rdQ3rmUzpAnnivJ+O305VPenwrnZY7WY+pE4H1xHk4sclf+rHJ6uwXX3zRjSfMns03v/nNlGI5GQV4YhbaiRlWSWWaTiV+boCvtXb7mCSfwyrcB3C4JRTecyiVQkqN0BLtpXcpBKbEeyI92FIAMbenm+NDieRzM902xsIrE1NeZZpOPk9QGl4ZtXKLGqT7UGID3CfaC4mpesqwHelWo7lPc/e6gGu3ZFkIjVZerw7pPa0zg/UZ+T8g6YFlFzh64kgf1wU03Wo+EVhPBrTpoJ0OQicDZK01W7dupaSkhJKSkt5N+ASU7SvD5/MxdOjQXt6F0j1Pw03ve3QywD8Z6Ke/90ktxNNB/ZOs6ZM13UtXOEmQTh/niQkdJ1r46f8nldTJrvHfAvgkcbene4M7IOE+YxWtUUKm8tpdwHc7/AiUqwTArcbDRmoB2vBaE3jcSqobkPvkd3Sv8Cc9z6j0AFxI90nMQrgtTd1uSt7jxVzgl6Kna53QjlepLN2HJSMyNE5G/s9a+kJ4tUqpvi29gSjduv2k4HbS+j2xA+tnAaATn8Hwcc9n+DSw/rzm5GTA+1lef5IiONlzKE4E8s96H/+ROfovA35GMpKRjPx3eR0nA67/iPLISAbwM5KRjGQkIyeRDLeRkYxkJCMZwM9IRjKSkYxkAD8jGclIRjKSAfyMZCQjGclIBvAzkpGMZCQjGcDPSEYykpGMZAA/IxnJSEYy8rnI/wcq6FLr0sFGSAAAAABJRU5ErkJggg=="

NAVY="0F172A";GOLD="B8972A";WHITE="FFFFFF";HDRBLUE="1E3A5F";GREEN="166634"
OFF="D1D5DB";OFFHDR="6B7280"
thin=Side(style="thin",color="CBD5E1")
bord=Border(left=thin,right=thin,top=thin,bottom=thin)
def fl(h): return PatternFill("solid",fgColor=h)
def al(h="center",v="center",w=False): return Alignment(horizontal=h,vertical=v,wrap_text=w)

def add_logo(ws,row=1):
    buf=io.BytesIO(base64.b64decode(LOGO_B64))
    img=XLImage(buf); img.width=280; img.height=32
    ws.add_image(img,f"A{row}")
    ws.row_dimensions[row].height=34

def day_name(year,month,d):
    return ["MON","TUE","WED","THU","FRI","SAT","SUN"][datetime.date(year,month,d).weekday()]

def is_off_day(year,month,d,off_day="sun"):
    dw=datetime.date(year,month,d).weekday()
    return dw==6 if off_day=="sun" else dw==4

def cc(ws,r,c,val="",bold=False,sz=9,color="000000",bg=None,ha="center",va="center",wrap=False):
    cell=ws.cell(row=r,column=c,value=val)
    cell.font=Font(name="Arial",bold=bold,size=sz,color=color)
    cell.alignment=al(ha,va,wrap)
    if bg: cell.fill=fl(bg)
    cell.border=bord
    return cell

def merge_cc(ws,r1,c1,r2,c2,val="",bold=False,sz=9,color="000000",bg=None,ha="center",va="center",wrap=False):
    ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
    c=ws.cell(row=r1,column=c1,value=val)
    c.font=Font(name="Arial",bold=bold,size=sz,color=color)
    c.alignment=al(ha,va,wrap)
    if bg: c.fill=fl(bg)
    for row in range(r1,r2+1):
        for col in range(c1,c2+1):
            ws.cell(row=row,column=col).border=bord
    return c

def print_setup(ws):
    ws.page_setup.orientation="landscape"
    ws.page_setup.paperSize=9
    ws.page_setup.fitToPage=True
    ws.page_setup.fitToWidth=1
    ws.page_setup.fitToHeight=0
    ws.print_title_rows="1:4"
    ws.page_margins=PageMargins(left=0.25,right=0.25,top=0.4,bottom=0.4,header=0.2,footer=0.2)
    ws.sheet_properties.pageSetUpPr.fitToPage=True

def sort_workers(workers):
    def key(w):
        parts=(w.get("emp","") or "").split("-")
        try: return (int(parts[1]),int(parts[2]))
        except: return (9999,0)
    return sorted(workers,key=key)

SKIP={"L.L.C","S.P.C","AND","GENERAL","CONTRACTING","LLC","SPC","-","&","MAINTENANCE","WORKS","CO.","COMPANY","SERVICES","TRADING","EST"}
def short_supplier(name,maxlen=18):
    words=[w for w in (name or "").upper().split() if w not in SKIP]
    r=""
    for word in words:
        nxt=(r+(" " if r else "")+word.capitalize())
        if len(nxt)<=maxlen: r=nxt
        else: break
    return r or (name or "")[:maxlen]

def build_customer(ws,data):
    year=data["year"]; month=data["month"]
    settings=data["settings"]; workers=sort_workers(data["workers"])
    holidays=data.get("holidays",{}); off_day=settings.get("weekendDay","sun")
    days_in_month=calendar.monthrange(year,month)[1]
    month_name=datetime.date(year,month,1).strftime("%B %Y").upper()
    def is_hol(d): return f"{year}-{month:02d}-{d:02d}" in holidays
    def off(d): return is_hol(d) or is_off_day(year,month,d,off_day)
    DAY_COL=5; TOT_COL=DAY_COL+days_in_month

    add_logo(ws,1)
    merge_cc(ws,1,5,1,TOT_COL,"SATCO ARABIA GENERAL CONTRACTING L.L.C - S.P.C\n"+settings.get("project",""),bold=True,sz=9,color=NAVY,bg="F8FAFC",ha="center",wrap=True)
    ws.row_dimensions[2].height=13
    merge_cc(ws,2,1,2,2,"CLIENT:",bold=True,sz=8,color=NAVY,bg="F8FAFC",ha="right")
    merge_cc(ws,2,3,2,DAY_COL+9,settings.get("clientName",""),bold=True,sz=8,color=NAVY,bg="F8FAFC",ha="left")
    merge_cc(ws,2,DAY_COL+10,2,DAY_COL+15,"MONTH: "+month_name,bold=True,sz=9,color=NAVY,bg="F8FAFC",ha="center")
    merge_cc(ws,2,DAY_COL+16,2,TOT_COL,settings.get("project",""),bold=True,sz=8,color=NAVY,bg="F8FAFC",ha="left")
    ws.row_dimensions[3].height=14; ws.row_dimensions[4].height=11
    for col,lbl in [(1,"SL\nNO"),(2,"EMP No."),(3,"NAME"),(4,"CRAFT")]:
        merge_cc(ws,3,col,4,col,lbl,bold=True,sz=8,color=WHITE,bg=HDRBLUE,ha="center",wrap=True)
    for d in range(1,days_in_month+1):
        col=DAY_COL+d-1; is_o=off(d)
        hbg=OFF if is_o else HDRBLUE; hfc=OFFHDR if is_o else WHITE
        c3=ws.cell(row=3,column=col,value=d)
        c3.font=Font(name="Arial",bold=True,size=7,color=hfc); c3.fill=fl(hbg); c3.alignment=al(); c3.border=bord
        c4=ws.cell(row=4,column=col,value=day_name(year,month,d))
        c4.font=Font(name="Arial",bold=True,size=7,color=hfc); c4.fill=fl(hbg); c4.alignment=al(); c4.border=bord
    merge_cc(ws,3,TOT_COL,4,TOT_COL,"TOTAL\nHRS",bold=True,sz=8,color=GOLD,bg=NAVY,ha="center",wrap=True)

    dr=5; grand=0
    for idx,w in enumerate(workers):
        total=sum(float(v) for v in w["hours"].values()); grand+=total
        rb="FFFFFF" if idx%2==0 else "F8FAFC"
        sn=" ".join(p.capitalize() for p in (w.get("name","") or "").split()[:2])
        cc(ws,dr,1,idx+1,bold=True,sz=8,bg=rb)
        cc(ws,dr,2,w.get("emp",""),sz=8,bg=rb,ha="left")
        cc(ws,dr,3,sn,bold=True,sz=8,bg=rb,ha="left")
        cc(ws,dr,4,w.get("craft",""),sz=8,bg=rb)
        for d in range(1,days_in_month+1):
            col=DAY_COL+d-1
            h=float(w["hours"].get(str(d),w["hours"].get(d,0)) or 0)
            is_o=off(d)
            c=ws.cell(row=dr,column=col,value=int(h) if h and h==int(h) else (h if h else ""))
            c.font=Font(name="Arial",size=8,bold=bool(h),color="0F172A" if h else "CBD5E1")
            c.fill=fl("E5E7EB" if is_o else rb); c.alignment=al(); c.border=bord
        tc=ws.cell(row=dr,column=TOT_COL,value=int(total) if total==int(total) else total)
        tc.font=Font(name="Arial",bold=True,size=9,color=NAVY); tc.fill=fl("DBEAFE"); tc.alignment=al(); tc.border=bord
        ws.row_dimensions[dr].height=14; dr+=1

    merge_cc(ws,dr,1,dr,4,"TOTAL HOURS",bold=True,sz=9,color=WHITE,bg=NAVY,ha="right")
    for d in range(1,days_in_month+1):
        col=DAY_COL+d-1
        dt=sum(float(w["hours"].get(str(d),w["hours"].get(d,0)) or 0) for w in workers)
        c=ws.cell(row=dr,column=col,value=int(dt) if dt and dt==int(dt) else (dt if dt else ""))
        c.font=Font(name="Arial",bold=True,size=8,color=GOLD); c.fill=fl(NAVY); c.alignment=al(); c.border=bord
    tc=ws.cell(row=dr,column=TOT_COL,value=int(grand) if grand==int(grand) else grand)
    tc.font=Font(name="Arial",bold=True,size=10,color=GOLD); tc.fill=fl(NAVY); tc.alignment=al(); tc.border=bord
    ws.row_dimensions[dr].height=16; dr+=2

    crafts={}
    for w in workers:
        cr=w.get("craft","Other"); crafts[cr]=crafts.get(cr,0)+sum(float(v) for v in w["hours"].values())
    merge_cc(ws,dr,1,dr,2,"CRAFT SUMMARY",bold=True,sz=8,color=WHITE,bg=NAVY)
    cc(ws,dr,3,"TOTAL HRS",bold=True,sz=8,color=WHITE,bg=NAVY); ws.row_dimensions[dr].height=13; dr+=1
    for craft,hrs in crafts.items():
        merge_cc(ws,dr,1,dr,2,craft,bold=True,sz=8,bg="F0F9FF",ha="left")
        cc(ws,dr,3,int(hrs) if hrs==int(hrs) else hrs,bold=True,sz=9,bg="DBEAFE"); ws.row_dimensions[dr].height=13; dr+=1
    merge_cc(ws,dr,1,dr,2,"TOTAL",bold=True,sz=9,color=WHITE,bg=NAVY,ha="right")
    tot=sum(crafts.values())
    cc(ws,dr,3,int(tot) if tot==int(tot) else tot,bold=True,sz=10,color=GOLD,bg=NAVY); ws.row_dimensions[dr].height=14; dr+=3

    third=TOT_COL//3
    for (sc,ec),txt,sub in zip([(1,third),(third+1,third*2),(third*2+1,TOT_COL)],
        ["PREPARED BY","CHECKED BY","APPROVED BY"],
        ["SATCO ARABIA GENERAL CONTRACTING LLC SPC",settings.get("checkedBy","Quality International"),settings.get("approvedBy","Alghanim International")]):
        merge_cc(ws,dr,sc,dr,ec,txt,bold=True,sz=9,color=NAVY,ha="center")
        ws.cell(row=dr,column=sc).font=Font(name="Arial",bold=True,size=9,color=NAVY,underline="single")
        ws.row_dimensions[dr].height=13
        merge_cc(ws,dr+1,sc,dr+1,ec,sub,sz=8,color="374151",ha="center"); ws.row_dimensions[dr+1].height=13

    ws.column_dimensions["A"].width=4; ws.column_dimensions["B"].width=12
    ws.column_dimensions["C"].width=16; ws.column_dimensions["D"].width=9
    for d in range(1,days_in_month+1): ws.column_dimensions[get_column_letter(DAY_COL+d-1)].width=3.0
    ws.column_dimensions[get_column_letter(TOT_COL)].width=8
    ws.freeze_panes=ws.cell(row=5,column=DAY_COL); print_setup(ws)

def build_supplier(ws,data,supp_label):
    year=data["year"]; month=data["month"]
    settings=data["settings"]; workers=sort_workers(data["workers"])
    holidays=data.get("holidays",{}); rates=data.get("rates",{})
    off_day=settings.get("weekendDay","sun")
    vat_pct=float(settings.get("vat","5"))/100
    def_rate=float(settings.get("defaultSupplierRate","16"))
    days_in_month=calendar.monthrange(year,month)[1]
    month_name=datetime.date(year,month,1).strftime("%B %Y").upper()
    def is_hol(d): return f"{year}-{month:02d}-{d:02d}" in holidays
    def off(d): return is_hol(d) or is_off_day(year,month,d,off_day)
    FIX=5; DS=FIX+1; FIN=DS+days_in_month; LAST=FIN+5

    add_logo(ws,1)
    merge_cc(ws,1,5,1,LAST,"SATCO ARABIA GENERAL CONTRACTING L.L.C - S.P.C\n"+settings.get("project",""),bold=True,sz=9,color=NAVY,bg="F8FAFC",ha="center",wrap=True)
    ws.row_dimensions[2].height=14
    merge_cc(ws,2,1,2,FIX,f"TIMESHEET - {month_name}",bold=True,sz=10,color=GOLD,bg=NAVY,ha="center")
    merge_cc(ws,2,FIX+1,2,LAST,f"Sub-Contractor: {supp_label}",bold=True,sz=9,color=NAVY,bg="FFFBEB",ha="center")
    ws.row_dimensions[3].height=14; ws.row_dimensions[4].height=11
    for ci,h in enumerate(["SL","EMP No","NAME","CRAFT","SUPPLIER"]):
        merge_cc(ws,3,ci+1,4,ci+1,h,bold=True,sz=8,color=WHITE,bg=GREEN)
    for d in range(1,days_in_month+1):
        col=DS+d-1; is_o=off(d)
        bg=OFF if is_o else GREEN; fc=OFFHDR if is_o else WHITE
        c3=ws.cell(row=3,column=col,value=d)
        c3.font=Font(name="Arial",bold=True,size=6,color=fc); c3.fill=fl(bg); c3.alignment=al(); c3.border=bord
        c4=ws.cell(row=4,column=col,value=day_name(year,month,d))
        c4.font=Font(name="Arial",bold=True,size=6,color=fc); c4.fill=fl(bg); c4.alignment=al(); c4.border=bord
    for fi,h in enumerate(["Total\nHrs","Rate","Deduct.","Gross Amt","VAT","Amount"]):
        merge_cc(ws,3,FIN+fi,4,FIN+fi,h,bold=True,sz=8,color=GOLD,bg=NAVY,ha="center",wrap=True)

    dr=5; gh=gg=gv=0
    for idx,w in enumerate(workers):
        total=sum(float(v) for v in w["hours"].values())
        craft=w.get("craft","Other")
        rate=float((rates.get(craft) or {}).get("supplierRate") or def_rate)
        gross=round(total*rate,2); vat=round(gross*vat_pct,2); amt=round(gross+vat,2)
        gh+=total; gg+=gross; gv+=vat
        rb="FFFFFF" if idx%2==0 else "F0FDF4"
        sn=" ".join(p.capitalize() for p in (w.get("name","") or "").split()[:2])
        supp_display=short_supplier(w.get("supplier","") or supp_label)
        cc(ws,dr,1,idx+1,bold=True,sz=8,bg=rb)
        cc(ws,dr,2,w.get("emp",""),sz=8,bg=rb,ha="left")
        cc(ws,dr,3,sn,bold=True,sz=8,bg=rb,ha="left")
        cc(ws,dr,4,craft,sz=8,bg=rb)
        c_s=ws.cell(row=dr,column=5,value=supp_display)
        c_s.font=Font(name="Arial",size=7,color="374151"); c_s.alignment=Alignment(horizontal="left",vertical="center",wrap_text=False); c_s.fill=fl(rb); c_s.border=bord
        for d in range(1,days_in_month+1):
            col=DS+d-1
            h_val=float(w["hours"].get(str(d),w["hours"].get(d,0)) or 0)
            c=ws.cell(row=dr,column=col,value=int(h_val) if h_val and h_val==int(h_val) else (h_val if h_val else ""))
            c.font=Font(name="Arial",size=8,bold=bool(h_val),color="0F172A" if h_val else "CBD5E1")
            c.fill=fl("E5E7EB" if off(d) else rb); c.alignment=al(); c.border=bord
        for fi,(v,fbg,fb,fmt) in enumerate(zip([total,rate,0,gross,vat,amt],["DBEAFE","FFFFFF","FFFFFF","F0FDF4","FFFFFF","F0FDF4"],[True,False,False,True,False,True],[None,None,None,"#,##0.00","#,##0.00","#,##0.00"])):
            c=ws.cell(row=dr,column=FIN+fi,value=v)
            c.font=Font(name="Arial",bold=fb,size=8,color="166634" if fb else "374151")
            c.fill=fl(fbg); c.alignment=al("right"); c.border=bord
            if fmt: c.number_format=fmt
        ws.row_dimensions[dr].height=14; dr+=1

    merge_cc(ws,dr,1,dr,FIX,"TOTAL",bold=True,sz=9,color=WHITE,bg=NAVY,ha="right")
    for d in range(1,days_in_month+1):
        col=DS+d-1
        dt=sum(float(w["hours"].get(str(d),w["hours"].get(d,0)) or 0) for w in workers)
        c=ws.cell(row=dr,column=col,value=int(dt) if dt and dt==int(dt) else (dt if dt else ""))
        c.font=Font(name="Arial",bold=True,size=8,color=GOLD); c.fill=fl(NAVY); c.alignment=al(); c.border=bord
    ga=round(gg+gv,2)
    for fi,(v,fmt) in enumerate([(round(gh,1),None),(None,None),(0,None),(round(gg,2),"#,##0.00"),(round(gv,2),"#,##0.00"),(ga,"#,##0.00")]):
        c=ws.cell(row=dr,column=FIN+fi,value=v if v is not None else "")
        c.font=Font(name="Arial",bold=True,size=9,color=GOLD); c.fill=fl(NAVY); c.alignment=al("right"); c.border=bord
        if fmt: c.number_format=fmt
    ws.row_dimensions[dr].height=16; dr+=2

    cm={}
    for w in workers:
        cr=w.get("craft","Other"); h2=sum(float(v) for v in w["hours"].values())
        r2=float((rates.get(cr) or {}).get("supplierRate") or def_rate)
        g2=round(h2*r2,2); v2=round(g2*vat_pct,2)
        if cr not in cm: cm[cr]={"h":0,"g":0,"v":0,"a":0}
        cm[cr]["h"]+=h2; cm[cr]["g"]+=g2; cm[cr]["v"]+=v2; cm[cr]["a"]+=round(g2+v2,2)
    for fi,h in enumerate(["CRAFT","TOTAL HRS","GROSS AMT","DEDUCT.","VAT","AMOUNT"]):
        c=ws.cell(row=dr,column=FIN+fi,value=h)
        c.font=Font(name="Arial",bold=True,size=8,color=WHITE); c.fill=fl(NAVY); c.alignment=al("center"); c.border=bord
    ws.row_dimensions[dr].height=13; dr+=1
    for craft,d2 in cm.items():
        for fi,(v,ha) in enumerate([(craft,"left"),(d2["h"],"right"),(d2["g"],"right"),(0,"right"),(d2["v"],"right"),(d2["a"],"right")]):
            c=ws.cell(row=dr,column=FIN+fi,value=v)
            c.font=Font(name="Arial",bold=(fi==0),size=8,color="166634"); c.fill=fl("F0FDF4"); c.alignment=al(ha); c.border=bord
            if fi>=2: c.number_format="#,##0.00"
        ws.row_dimensions[dr].height=13; dr+=1
    gh2=sum(d["h"] for d in cm.values()); gg2=sum(d["g"] for d in cm.values())
    gv2=sum(d["v"] for d in cm.values()); ga2=sum(d["a"] for d in cm.values())
    for fi,(v,ha,fmt) in enumerate([("GRAND TOTAL","left",None),(gh2,"right",None),(round(gg2,2),"right","#,##0.00"),(0,"right",None),(round(gv2,2),"right","#,##0.00"),(round(ga2,2),"right","#,##0.00")]):
        c=ws.cell(row=dr,column=FIN+fi,value=v)
        c.font=Font(name="Arial",bold=True,size=9,color=GOLD); c.fill=fl(NAVY); c.alignment=al(ha); c.border=bord
        if fmt: c.number_format=fmt
    ws.row_dimensions[dr].height=15; dr+=3

    third=LAST//3
    for (sc,ec),txt,sub in zip([(1,third),(third+1,third*2),(third*2+1,LAST)],
        ["PREPARED BY","CHECKED BY","APPROVED BY"],
        ["SATCO ARABIA GENERAL CONTRACTING LLC SPC",settings.get("checkedBy","Quality International"),settings.get("approvedBy","Alghanim International")]):
        merge_cc(ws,dr,sc,dr,ec,txt,bold=True,sz=9,color=NAVY,ha="center")
        ws.cell(row=dr,column=sc).font=Font(name="Arial",bold=True,size=9,color=NAVY,underline="single")
        ws.row_dimensions[dr].height=13
        merge_cc(ws,dr+1,sc,dr+1,ec,sub,sz=8,color="374151",ha="center"); ws.row_dimensions[dr+1].height=13

    ws.column_dimensions["A"].width=3.5; ws.column_dimensions["B"].width=12
    ws.column_dimensions["C"].width=14; ws.column_dimensions["D"].width=7.5
    ws.column_dimensions["E"].width=14
    for d in range(1,days_in_month+1): ws.column_dimensions[get_column_letter(DS+d-1)].width=2.8
    for fi,w in enumerate([9,6,9,13,9,13]): ws.column_dimensions[get_column_letter(FIN+fi)].width=w
    ws.freeze_panes=ws.cell(row=5,column=DS); print_setup(ws)

class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200); self._cors(); self.end_headers()
    def _cors(self):
        self.send_header("Access-Control-Allow-Origin","*")
        self.send_header("Access-Control-Allow-Methods","POST,OPTIONS")
        self.send_header("Access-Control-Allow-Headers","Content-Type")
    def do_POST(self):
        try:
            length=int(self.headers.get("Content-Length",0))
            body=json.loads(self.rfile.read(length))
            ts_type=body.get("type","customer")
            data=body.get("data",{})
            supp=body.get("supplier",None)
            fname=body.get("filename","SATCO_Timesheet.xlsx")
            wb=openpyxl.Workbook(); wb.remove(wb.active)
            if ts_type=="customer":
                ws=wb.create_sheet("Timesheet"); build_customer(ws,data)
            else:
                ws_all=wb.create_sheet("All Suppliers"); build_supplier(ws_all,data,supp or "ALL SUPPLIERS")
                if not supp:
                    sm=defaultdict(list)
                    for w in data.get("workers",[]): sm[w.get("supplier","")].append(w)
                    def sk(item):
                        sw=sort_workers(item[1])
                        p=(sw[0]["emp"] if sw else "").split("-")
                        try: return (int(p[1]),int(p[2]))
                        except: return (9999,0)
                    for sn,sw in sorted(sm.items(),key=sk):
                        sub={**data,"workers":sw}
                        ws2=wb.create_sheet(sn[:31]); build_supplier(ws2,sub,sn)
            buf=io.BytesIO(); wb.save(buf); xlsx=buf.getvalue()
            self.send_response(200); self._cors()
            self.send_header("Content-Type","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition",f'attachment; filename="{fname}"')
            self.send_header("Content-Length",str(len(xlsx))); self.end_headers()
            self.wfile.write(xlsx)
        except Exception as e:
            err=str(e).encode()
            self.send_response(500); self._cors()
            self.send_header("Content-Type","text/plain")
            self.send_header("Content-Length",str(len(err))); self.end_headers()
            self.wfile.write(err)
    def log_message(self,*a): pass
